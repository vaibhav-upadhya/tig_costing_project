from datetime import datetime
from django.shortcuts import render
from django.shortcuts import render, redirect , get_object_or_404
from .models import ExpenseHead, ExpenseField, Costing, Department, AuditTrail
from .forms import CostingForm
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from tig_costing.models import ExpenseHead, ExpenseField
from .models import ExpenseField
from .models import ExpenseHead
from django.http import JsonResponse
from django.urls import reverse
from collections import defaultdict
from collections import defaultdict
from collections import defaultdict
from django.shortcuts import render
from .models import ExpenseHead, Department, ExpenseField, Costing
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from tig_costing.models import ExpenseField, ExpenseType
import openpyxl
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook







def home(request):
    return render(request, 'tig_costing/home.html')




# AJAX view to load Expense Fields dynamically based on selected Expense Head
def load_expense_fields(request):
    expense_head_id = request.GET.get('expense_head')
    expense_fields = ExpenseField.objects.filter(expense_head__id=expense_head_id)
    data = {"expense_fields": [{"id": field.id, "name": field.name} for field in expense_fields]}
    print("load_expense_fields view was called")

    return JsonResponse(data)

def load_expense_heads(request):
    department_id = request.GET.get('department')
    if not department_id:
        return JsonResponse({"error": "Department ID not provided"}, status=400)

    try:
        expense_heads = ExpenseHead.objects.filter(expensefield__department__id=department_id).distinct()
    except Department.DoesNotExist:
        return JsonResponse({"error": "Invalid department ID"}, status=400)

    data = {"expense_heads": [{"id": head.id, "name": head.name} for head in expense_heads]}
    return JsonResponse(data)






# @login_required
# def add_costing(request, expense_head_id=None):
#     departments = Department.objects.all()
#     expense_heads = ExpenseHead.objects.all()  # Fetch all Expense Heads
#     expense_fields = ExpenseField.objects.all()

#     if request.method == "POST":
        
#         form = CostingForm(request.POST)
#         if form.is_valid():
#             department = form.cleaned_data['department']
#             expense_head = form.cleaned_data['expense_head']
#             expense_field = form.cleaned_data['expense_field']
#             month = form.cleaned_data['month']
#             year = form.cleaned_data['year']
#             cost = form.cleaned_data['cost']

#             # Save the data to the Costing model
#             Costing.objects.create(
#                 department=department,
#                 expense_head=expense_head,
#                 expense_field=expense_field,
#                 month=month,
#                 year=year,
#                 cost=cost,
#                 updated_by=request.user
#             )
#             return redirect('view_costing', expense_head_id=expense_head.id)
#     else:
#         # For GET request, initialize an empty form with dynamic fields
#         department_id = request.GET.get('department')  # Get the department id from query parameters (optional)
#         form = CostingForm(department_id=department_id)

#         # Dynamically populate expense fields and heads based on department
#         if department_id:
#             expense_fields = ExpenseField.objects.filter(department__id=department_id)
#             expense_heads = ExpenseHead.objects.filter(expensefield__department__id=department_id).distinct()

#         # Debugging: Check if expense_fields and expense_heads are being correctly populated
#         # print("Expense Heads:", expense_heads)
#         # print("Expense Fields:", expense_fields)
        

#     return render(request, 'tig_costing/add_costing.html', {
#         'form': form,
#         'departments': departments,
#         'expense_heads': expense_heads,
#         'expense_fields': expense_fields
#     })

#=========== testing new functionality======================================

def costing_view(request):
    if request.method == "POST":
        form = CostingForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('success')  # Redirect to a success page after saving the form
    else:
        form = CostingForm()

    departments = Department.objects.all()  # To render departments in a dropdown
    return render(request, 'test.html', {'form': form, 'departments': departments})








@login_required
def add_costing(request, expense_head_id=None):
    departments = Department.objects.all()
    expense_heads = ExpenseHead.objects.all()
    expense_fields = ExpenseField.objects.all()

    if request.method == "POST":
        form = CostingForm(request.POST)
        if form.is_valid():
            department = form.cleaned_data['department']
            expense_head = form.cleaned_data['expense_head']
            expense_field = form.cleaned_data['expense_field']
            month = form.cleaned_data['month']
            year = form.cleaned_data['year']
            cost = form.cleaned_data['cost']

            # Build the hierarchical JSON format
            data = {
                expense_head.name: {
                    department.name: {
                        expense_field.name: {
                            f"{month}'{year % 100}": float(cost)
                        }
                    }
                }
            }

            # Check if an entry already exists for the given fields
            costing, created = Costing.objects.get_or_create(
                department=department,
                expense_head=expense_head,
                expense_field=expense_field,
                year=year,
                month=month,
                defaults={
                    'cost': cost,
                    'data': data,
                    'updated_by': request.user,
                }
            )

            if not created:
                # Update existing entry and merge JSON data
                existing_data = costing.data or {}
                try:
                    # Merge the new data with existing data
                    existing_data = merge_costing_data(existing_data, data)
                    costing.data = existing_data
                    costing.cost = cost
                    costing.updated_by = request.user
                    costing.save()
                except ValidationError as e:
                    return JsonResponse({"error": str(e)}, status=400)

            return redirect('view_costing', expense_head_id=expense_head.id)

    else:
        department_id = request.GET.get('department')
        form = CostingForm(department_id=department_id)
        if department_id:
            expense_fields = ExpenseField.objects.filter(department__id=department_id)
            expense_heads = ExpenseHead.objects.filter(expensefield__department__id=department_id).distinct()

    return render(request, 'tig_costing/add_costing.html', {
        'form': form,
        'departments': departments,
        'expense_heads': expense_heads,
        'expense_fields': expense_fields
    })


def merge_costing_data(existing_data, new_data):
    """Merge existing data with new data."""
    for key, value in new_data.items():
        if key in existing_data:
            if isinstance(value, dict) and isinstance(existing_data[key], dict):
                existing_data[key] = merge_costing_data(existing_data[key], value)
            else:
                raise ValidationError("Conflicting data structure.")
        else:
            existing_data[key] = value
    return existing_data


@login_required
def view_costing(request, expense_head_id):
    try:
        expense_head = ExpenseHead.objects.get(id=expense_head_id)
    except ExpenseHead.DoesNotExist:
        return render(request, 'tig_costing/error.html', {'message': 'Expense Head not found'})

    costings = Costing.objects.filter(expense_head=expense_head)

    # Aggregate hierarchical data
    expenses = {}
    for costing in costings:
        data = costing.data or {}
        expenses = merge_costing_data(expenses, data)

    context = {
        'expense_head': expense_head,
        'expenses': expenses,
    }
    return render(request, 'tig_costing/view_costing.html', context)



@login_required
def get_costing_data(request, expense_head_id):
    try:
        expense_head = ExpenseHead.objects.get(id=expense_head_id)
    except ExpenseHead.DoesNotExist:
        return JsonResponse({"error": "Expense Head not found"}, status=404)

    costings = Costing.objects.filter(expense_head=expense_head)
    data = {}
    for costing in costings:
        data = merge_costing_data(data, costing.data or {})

    return JsonResponse(data)



@csrf_exempt
def add_expense_type(request):
    if request.method == "POST":
        data = request.POST
        expense_field_id = data.get("expense_field_id")
        expense_name = data.get("expense_name")

        if not expense_field_id or not expense_name:
            return JsonResponse({"error": "Missing expense_field_id or expense_name"}, status=400)

        try:
            expense_field = ExpenseField.objects.get(id=expense_field_id)
        except ExpenseField.DoesNotExist:
            return JsonResponse({"error": "ExpenseField not found"}, status=404)

        # Create the ExpenseType (signal will handle costing creation)
        expense_type = ExpenseType.objects.create(
            expense_field=expense_field,
            expense_name=expense_name
        )

        return JsonResponse({
            "message": "ExpenseType and associated Costing records created successfully.",
            "expense_type": {
                "id": str(expense_type.id),
                "name": expense_type.expense_name,
                "expense_field": expense_field.name
            }
        })

    return JsonResponse({"error": "Invalid request method."}, status=405)





@user_passes_test(lambda user: user.is_superuser)
def create_dynamic_field(request):
    """Allow superusers to dynamically create departments or expense fields."""
    if request.method == "POST":
        department_name = request.POST.get("department_name")
        expense_field_name = request.POST.get("expense_field_name")
        expense_head_id = request.POST.get("expense_head_id")

        # Create department if provided
        if department_name:
            department, _ = Department.objects.get_or_create(name=department_name, is_dynamic=True)

        # Create expense field if provided
        if expense_field_name and expense_head_id:
            expense_head = get_object_or_404(ExpenseHead, id=expense_head_id)
            department = get_object_or_404(Department, name=department_name)
            ExpenseField.objects.create(
                name=expense_field_name,
                expense_head=expense_head,
                department=department,
            )
        return JsonResponse({"status": "success"})

    return JsonResponse({"error": "Invalid request"}, status=400)

def test_view(request):
    # Fetch all expense heads
    expense_heads = ExpenseHead.objects.all()

    # Create an empty dictionary to store all the hierarchical data
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))

    # Fetch the costings and populate the data
    for expense_head in expense_heads:
        costings = Costing.objects.filter(expense_head=expense_head)

        for costing in costings:
            month_key = costing.month[:3] + "'25"  # Example: Jan'25
            department = costing.department.name
            expense_field = costing.expense_field.name
            year_key = str(costing.year)

            # Insert the costing data into the hierarchical structure
            data[expense_head.name][department][expense_field][f"{month_key}"] = str(costing.cost)
            print(data)

    # Pass the data to the template
    return render(request, 'tig_costing/test.html', {'expense_data': data})


def generate_nested_costing_structure(request):
    costing_structure = {}
    expense_fields = ExpenseField.objects.select_related('expense_head', 'department').all().order_by('name')
    expense_types = ExpenseType.objects.prefetch_related('costing_set')
 
    expense_head_departments = {}
    year = datetime.now().year
 
    for expense_field in expense_fields:
        expense_head = getattr(expense_field.expense_head, 'name', None)
        department = getattr(expense_field.department, 'name', None)
 
        if not expense_head or not department:
            continue
 
        if expense_head not in costing_structure:
            costing_structure[expense_head] = {
                "multiple_departments": False,
            }
            expense_head_departments[expense_head] = set()
 
        if department not in costing_structure[expense_head]:
            costing_structure[expense_head][department] = {}
 
        expense_head_departments[expense_head].add(department)
 
        related_expense_types = expense_types.filter(expense_field=expense_field)
        for expense_type in related_expense_types:
            expense_name = expense_type.expense_name
            if expense_name not in costing_structure[expense_head][department]:
                costing_structure[expense_head][department][expense_name] = {}
 
            costings = Costing.objects.filter(expense_type=expense_type, year = year)
            for costing in costings:
                month_year = f"{costing.month}'{str(costing.year)[-2:]}"
                costing_structure[expense_head][department][expense_name][month_year] = float(costing.cost or 0.0)
 
    for expense_head, departments in expense_head_departments.items():
        if len(departments) > 1:
            costing_structure[expense_head]["multiple_departments"] = True
 
    
    year =str(year)[2:]
   
    
    months = [
    f"Jan'{year}", f"Feb'{year}", f"Mar'{year}", f"Apr'{year}", f"May'{year}",
    f"Jun'{year}", f"Jul'{year}", f"Aug'{year}", f"Sept'{year}",
    f"Oct'{year}", f"Nov'{year}", f"Dec'{year}"
    ]
 
    context = {
        "costing_structure": costing_structure,
        "months": months
    }
    print(f"Nested Costing Structure: {len(costing_structure.keys())} {'*'*100}\n {context} expense heads processed.")
 
 
    return render(request, 'tig_costing/test.html', context)



def generate_excel(request):
    # Fetch all costings
    costings = Costing.objects.all()

    # Prepare data structure with nested defaultdict
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: Decimal('0')))))  # Added ExpenseType

    months_years_set = set()

    # Iterate over costings to populate the data structure
    for costing in costings:
        expense_type_str = costing.expense_type.expense_name
        
        # Access the department through the related ExpenseField
        department = costing.expense_type.expense_field.department.name  # Assuming department has a 'name' field
        
        # Extract the expense head (split the name by " - " if it exists)
        if ' - ' in expense_type_str:
            expense_head, _ = expense_type_str.split(' - ', 1)
        else:
            expense_head = expense_type_str

        # Access the expense type name
        expense_type_name = costing.expense_type.expense_name
        
        # Format month and year for sorting
        month = costing.month
        year = str(costing.year)[-2:]  # Extract the last two digits of the year
        month_year = f"{month}-{year}"
        months_years_set.add(month_year)

        # Add cost to data structure
        cost = Decimal(str(costing.cost))  # Convert to Decimal for calculations
        data[expense_head][department][expense_type_name][month_year] += cost

    # Sort months and years based on datetime formatting
    months_years_sorted = sorted(
        months_years_set, 
        key=lambda x: datetime.strptime(x.replace('Sept', 'Sep'), "%b-%y")
    )

    # Create an Excel workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Expense Sheet"

    # Add the "IN USD" header
    sheet.merge_cells('N1:O1')
    cell = sheet.cell(row=1, column=14, value='IN USD')
    cell.alignment = Alignment(horizontal='right')
    cell.font = Font(bold=True)

    # Add table headers
    headers = ['Expense Head', 'Department', 'Expense Type'] + months_years_sorted + ['Total']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")  # White font for better contrast
        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Populate the data into the table
    current_row = 3

    for expense_head, departments in data.items():
        # Write the parent category (expense head)
        sheet.cell(row=current_row, column=1, value=expense_head)
        current_row += 1

        for department, expense_types in departments.items():
            # Write department name
            sheet.cell(row=current_row, column=2, value=department)
            current_row += 1

            for expense_type_name, month_data in expense_types.items():
                # Write expense type name
                row = ['', '', expense_type_name]
                total_cost = Decimal('0')

                # Populate month-wise data
                for month_year in months_years_sorted:
                    cost = month_data.get(month_year, Decimal('0'))
                    row.append(cost)
                    total_cost += cost

                row.append(total_cost)  # Add total cost at the end

                # Write the row into the sheet
                for col, value in enumerate(row, 1):
                    cell = sheet.cell(row=current_row, column=col, value=value)
                    if col > 2:  # Format month and total columns
                        cell.number_format = '#,##0.00'
                current_row += 1

    # Apply table formatting
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

    # Set column widths
    sheet.column_dimensions['A'].width = 25  # Expense Head
    sheet.column_dimensions['B'].width = 20  # Department
    sheet.column_dimensions['C'].width = 25  # Expense Type
    for col in range(4, 4 + len(months_years_sorted)):
        sheet.column_dimensions[get_column_letter(col)].width = 12

    # Freeze panes
    sheet.freeze_panes = 'D3'

    # Prepare the HTTP response for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expense_sheet.xlsx'
    wb.save(response)
    return response